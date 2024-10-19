# From your terminal, run the following command:
# pip install pandas python-docx openpyxl openai python-dotenv

# Ensure the directory from which you're running this has a .env file with the following:
# OPENAI_API_KEY=<your key here>

import os
import pandas as pd
import docx
import csv
import openai
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Set up OpenAI API key
openai.api_key = os.getenv("OPENAI_API_KEY")

def translate_text(text, target_language):
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": f"You are a professional translator. Translate the following text to {target_language} in a professional business tone, as a native speaker would write it."},
            {"role": "user", "content": text}
        ]
    )
    return response.choices[0].message['content'].strip()

def translate_excel(file_path, target_language):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = translate_text(str(cell.value), target_language)
    output_path = f"{os.path.splitext(file_path)[0]}_{target_language}.xlsx"
    wb.save(output_path)
    return output_path

def translate_csv(file_path, target_language):
    df = pd.read_csv(file_path)
    for column in df.columns:
        df[column] = df[column].apply(lambda x: translate_text(str(x), target_language))
    output_path = f"{os.path.splitext(file_path)[0]}_{target_language}.csv"
    df.to_csv(output_path, index=False)
    return output_path

def translate_docx(file_path, target_language):
    doc = docx.Document(file_path)
    for para in doc.paragraphs:
        para.text = translate_text(para.text, target_language)
    output_path = f"{os.path.splitext(file_path)[0]}_{target_language}.docx"
    doc.save(output_path)
    return output_path

def main():
    file_path = input("Enter the path to the document: ").strip('"')  # Remove any surrounding quotes
    target_language = input("Enter the target language: ")

    print(f"File path: {file_path}")  # Debug print
    print(f"Target language: {target_language}")  # Debug print

    _, file_extension = os.path.splitext(file_path)
    print(f"Detected file extension: {file_extension}")  # Debug print

    if file_extension.lower() == '.xlsx':
        output_path = translate_excel(file_path, target_language)
    elif file_extension.lower() == '.csv':
        output_path = translate_csv(file_path, target_language)
    elif file_extension.lower() == '.docx':
        output_path = translate_docx(file_path, target_language)
    else:
        print(f"Unsupported file format: {file_extension}")
        return

    print(f"Translation complete. Output file: {output_path}")

if __name__ == "__main__":
    main()
